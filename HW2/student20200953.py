#!/usr/bin/python3
import openpyxl
import math

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']
scoreList = []

row_id = 1;
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value 
		ws.cell(row = row_id, column = 7).value = sum_v
		scoreList.append(sum_v)
	row_id += 1

scoreList.sort(reverse=True)
cnt = ws.max_row - 1


aCut = scoreList[math.floor(cnt * 0.3)]
apCut = scoreList[math.floor(cnt * 0.15)]
bCut = scoreList[math.floor(cnt * 0.7)]
bpCut = scoreList[math.floor(cnt * 0.5)]
cpCut = scoreList[math.floor(cnt * 0.85)]
row_id = 1
ap = 0
a = 0
bp = 0
b = 0
cp = 0
c = 0

for i in range(2, ws.max_row + 1):
	if apCut < ws.cell(row = i, column = 7).value:
		ap = ap + 1
	elif aCut < ws.cell(row = i, column = 7).value:
		a = a + 1
	elif bpCut < ws.cell(row = i, column = 7).value:
		bp = bp + 1
	elif bCut < ws.cell(row = i, column = 7).value:
		b = b + 1
	elif cpCut < ws.cell(row = i, column = 7).value:
		cp = cp + 1
	else:
		c = c + 1 

aCnt = math.floor(cnt * 0.3)
apCnt = math.floor(cnt * 0.15)
bCnt = math.floor(cnt * 0.7)
bpCnt = math.floor(cnt * 0.5)
cpCnt = math.floor(cnt * 0.85)

if ap > apCnt:
	ap = ap - scoreList.count(apCut)
	a = a + scoreList.count(apCut)
if a > aCnt:
	a = a - scoreList.count(aCut)
	bp = bp + scoreList.count(aCut)
if bp > bpCnt:
	bp = bp - scoreList.count(bpCut)
	b = b + scoreList.count(bpCut)
if b > bCnt:
	b = b - scoreList.count(bCut)
	cp = cp + scoreList.ount(bCut)
if cp > cpCnt:
	cp = cp - socreList.count(cpCut)
	c = c + scoreList.count(cpCut)

		 

for i in range(0, ap):
	for j in range(1, ws.max_row + 1):
		if scoreList[i] == ws.cell(row = j, column = 7).value:	
			ws.cell(row = j, column = 8).value = "A+"

for i in range(ap, a):
	for j in range(1, ws.max_row + 1):
		if scoreList[i] == ws.cell(row = j, column = 7).value:
			ws.cell(row = j, column = 8).value = "A" 

for i in range(a, bp):
	for j in range(1, ws.max_row + 1):
		if scoreList[i] == ws.cell(row = j, column = 7).value:
			ws.cell(row = j, column = 8).value = "B+"

for i in range(bp, b):
	for j in range(1, ws.max_row + 1):
		if scoreList[i] == ws.cell(row = j, column = 7).value:
			ws.cell(row = j, column = 8).value = "B"

for i in range(b, cp):
	for j in range(1, ws.max_row + 1):
		if scoreList[i] == ws.cell(row = j, column = 7).value:
			ws.cell(row = j, column = 8).value = "C+"

for i in range(cp, cnt):
	for j in range(1, ws.max_row + 1):
		if scoreList[i] == ws.cell(row = j, column = 7).value:
			ws.cell(row = j, column = 8).value = "C"
wb.save( "student.xlsx" )

